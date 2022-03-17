import openpyxl

wk = openpyxl.load_workbook("D:\\Downloads\\WorkSpace\\TestData\\test.xlsx")


def fetch_rows(sheet_name):
    sh = wk["Sheet1"]
    return sh.max_row


def fetch_cell_data(sheet_name, row, column):
    sh = wk["Sheet1"]
    return sh.cell(row=int(row), column=int(column)).value


